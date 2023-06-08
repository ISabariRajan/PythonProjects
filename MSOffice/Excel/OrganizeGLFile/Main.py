from openpyxl import Workbook, load_workbook

filename = "GLAccountUpdate.xlsx"
wb = load_workbook(filename)
output_sheet = "Output"
print(f"{'Output' in wb}")
if "Output" in wb:
    print("OUTPUT is there")
    del wb["Output"]
    wb.save(filename)

# print(f"Creating new sheet for output")
# wb.create_sheet(title="Output")
# work_sheet = wb["Output"]
# work_sheet.cell(row=2, column=2).value = "NAME"
# work_sheet.cell(row=2, column=3).value = "LEVEL 1"
# work_sheet.cell(row=2, column=4).value = "LEVEL 2"
# work_sheet.cell(row=2, column=5).value = "LEVEL 3"
# work_sheet.cell(row=2, column=6).value = "LEVEL 4"
# work_sheet.cell(row=2, column=7).value = "LEVEL 5"
# wb.save(filename)

sheet1 = wb["Level 1-4"]
sheet2 = wb["Level 5"]
sheet1_array = []
sheet2_array = []
# print("Capturing Data from ")
# print(f"Sheet: {sheet1}, Columns: {sheet1.max_column}, Rows: {sheet1.max_row}")
# for row1 in range(4, sheet1.max_row + 1):
#     l1 = sheet1.cell(row=row1, column=2).value
#     l2 = sheet1.cell(row=row1, column=3).value
#     l3 = sheet1.cell(row=row1, column=4).value
#     l4 = sheet1.cell(row=row1, column=5).value
#     sheet1_array.append((l1, l2, l3, l4))

# print("Capturing Data from ")
# print(f"Sheet: {sheet2}, Columns: {sheet2.max_column}, Rows: {sheet2.max_row}")
# for row2 in range(4, sheet2.max_row + 1):
#     l5 = sheet2.cell(row=row2, column=2).value
#     item = sheet2.cell(row=row2, column=3).value
#     sheet2_array.append((l5, item))


start_row = 4
# for sh1 in sheet1_array:
#     for sh2 in sheet2_array:
#         print(f"{sh2[1]} {sh1[0]} {sh1[1]} {sh1[2]} {sh1[3]} {sh2[0]}")
wb.save(filename)